"""
Vista de reportes - VERSIÓN SIN PANDAS
"""
import customtkinter as ctk
from tkinter import messagebox, filedialog
from datetime import datetime, timedelta
from .widgets import DateEntry
from models.expediente import Expediente
from models.prestamo import Prestamo
from models.historial import Historial
from fpdf import FPDF
import os
import csv
import openpyxl
from openpyxl import Workbook

class ReportePDF(FPDF):
    """PDF para reportes"""
    def header(self):
        self.set_font("Arial", 'B', 15)
        self.set_text_color(31, 106, 165)
        self.cell(0, 10, "SISTEMA DE EXPEDIENTES CLÍNICOS", 0, 1, 'C')
        self.set_font("Arial", 'I', 10)
        self.set_text_color(100)
        self.cell(0, 5, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", 0, 1, 'C')
        self.ln(5)

class ReportesView:
    """Vista para generar reportes - SIN PANDAS"""
    
    def __init__(self, parent, usuario_actual):
        self.parent = parent
        self.usuario_actual = usuario_actual
        self.frame = None
        self.crear_interfaz()
    
    def crear_interfaz(self):
        """Crea la interfaz de reportes"""
        self.frame = ctk.CTkFrame(self.parent, fg_color="transparent")
        self.frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        ctk.CTkLabel(
            self.frame,
            text="📊 REPORTES Y ESTADÍSTICAS",
            font=("Segoe UI", 24, "bold"),
            text_color="#1F6AA5"
        ).pack(pady=(10, 20))
        
        # Estadísticas
        stats_frame = ctk.CTkFrame(self.frame, fg_color="white", corner_radius=12)
        stats_frame.pack(pady=10, padx=20, fill="x")
        
        self.mostrar_estadisticas(stats_frame)
        
        # Opciones de reportes
        rep_frame = ctk.CTkFrame(self.frame, fg_color="transparent")
        rep_frame.pack(pady=20)
        
        ctk.CTkLabel(rep_frame, text="Seleccione el tipo de reporte:", font=("Segoe UI", 14, "bold")).pack(pady=10)
        
        self.reporte_tipo = ctk.CTkComboBox(
            rep_frame,
            values=[
                "Expedientes prestados",
                "Expedientes disponibles",
                "Expedientes vencidos",
                "Préstamos por mes",
                "Áreas que solicitan más expedientes"
            ],
            width=300,
            height=40
        )
        self.reporte_tipo.pack(pady=5)
        self.reporte_tipo.set("Expedientes prestados")
        
        # Rango de fechas
        fecha_frame = ctk.CTkFrame(rep_frame, fg_color="transparent")
        fecha_frame.pack(pady=10)
        
        ctk.CTkLabel(fecha_frame, text="Desde:", font=("Segoe UI", 12)).pack(side="left", padx=5)
        self.desde = DateEntry(fecha_frame, placeholder_text="DD/MM/AAAA", width=100, height=35)
        self.desde.pack(side="left", padx=5)
        self.desde.set_date((datetime.now() - timedelta(days=30)).strftime("%d/%m/%Y"))
        
        ctk.CTkLabel(fecha_frame, text="Hasta:", font=("Segoe UI", 12)).pack(side="left", padx=5)
        self.hasta = DateEntry(fecha_frame, placeholder_text="DD/MM/AAAA", width=100, height=35)
        self.hasta.pack(side="left", padx=5)
        self.hasta.set_date(datetime.now().strftime("%d/%m/%Y"))
        
        # Botones
        btn_frame = ctk.CTkFrame(rep_frame, fg_color="transparent")
        btn_frame.pack(pady=15)
        
        ctk.CTkButton(
            btn_frame,
            text="📄 Generar Reporte",
            command=self.generar_reporte,
            fg_color="#3498DB",
            hover_color="#2980B9",
            width=180,
            height=45,
            font=("Segoe UI", 14, "bold")
        ).pack(side="left", padx=10)
        
        ctk.CTkButton(
            btn_frame,
            text="📥 Exportar Excel",
            command=self.exportar_excel_sin_pandas,  # <--- NUEVO método
            fg_color="#27AE60",
            hover_color="#219150",
            width=180,
            height=45,
            font=("Segoe UI", 14, "bold")
        ).pack(side="left", padx=10)
        
        ctk.CTkButton(
            btn_frame,
            text="📄 Exportar PDF",
            command=self.exportar_pdf_sin_pandas,  # <--- NUEVO método
            fg_color="#2980B9",
            hover_color="#2471A3",
            width=180,
            height=45,
            font=("Segoe UI", 14, "bold")
        ).pack(side="left", padx=10)
        
        # Frame para resultados
        self.resultados_frame = ctk.CTkScrollableFrame(
            self.frame,
            fg_color="white",
            corner_radius=10,
            border_width=1
        )
        self.resultados_frame.pack(fill="both", expand=True, pady=10, padx=10)
    
    def mostrar_estadisticas(self, parent):
        """Muestra estadísticas generales"""
        stats = self._obtener_estadisticas()
        
        grid = ctk.CTkFrame(parent, fg_color="transparent")
        grid.pack(pady=10)
        
        stats_text = [
            (f"📋 Total: {stats['total']}", "#2C3E50"),
            (f"🟢 Disponibles: {stats['disponibles']}", "#27AE60"),
            (f"🟡 Prestados: {stats['prestados']}", "#F39C12"),
            (f"🔴 Archivo: {stats['archivo_muerto']}", "#E74C3C"),
            (f"⏰ Vencidos: {stats['vencidos']}", "#E67E22"),
            (f"📤 Préstamos activos: {stats['prestamos_activos']}", "#3498DB")
        ]
        
        for i, (texto, color) in enumerate(stats_text):
            row = i // 3
            col = i % 3
            ctk.CTkLabel(
                grid,
                text=texto,
                font=("Segoe UI", 13),
                text_color=color
            ).grid(row=row, column=col, padx=20, pady=5, sticky="w")
    
    def _obtener_estadisticas(self):
        """Obtiene estadísticas del sistema"""
        stats = {
            "total": 0, "disponibles": 0, "prestados": 0,
            "archivo_muerto": 0, "vencidos": 0, "prestamos_activos": 0
        }
        
        # Contar expedientes
        todos = Expediente.obtener_todos("Todos")
        stats["total"] = len(todos)
        
        # Contar por estado
        disponibles = Expediente.obtener_todos("Disponible")
        stats["disponibles"] = len(disponibles)
        
        prestados = Expediente.obtener_todos("Prestado")
        stats["prestados"] = len(prestados)
        
        archivados = Expediente.obtener_todos("Archivo muerto")
        stats["archivo_muerto"] = len(archivados)
        
        # Préstamos activos
        activos = Prestamo.obtener_todos_activos()
        stats["prestamos_activos"] = len(activos)
        
        # Vencidos
        vencidos = Prestamo.obtener_vencidos()
        stats["vencidos"] = len(vencidos)
        
        return stats
    
    def generar_reporte(self):
        """Genera el reporte seleccionado"""
        tipo = self.reporte_tipo.get()
        
        # Limpiar resultados
        for widget in self.resultados_frame.winfo_children():
            widget.destroy()
        
        if tipo == "Expedientes prestados":
            self._reporte_prestados()
        elif tipo == "Expedientes disponibles":
            self._reporte_disponibles()
        elif tipo == "Expedientes vencidos":
            self._reporte_vencidos()
        elif tipo == "Préstamos por mes":
            self._reporte_prestamos_mes()
        elif tipo == "Áreas que solicitan más expedientes":
            self._reporte_areas()
    
    def _reporte_prestados(self):
        """Reporte de expedientes prestados"""
        resultados = Prestamo.obtener_todos_activos()
        
        if not resultados:
            ctk.CTkLabel(
                self.resultados_frame,
                text="❌ No hay expedientes prestados actualmente.",
                font=("Segoe UI", 14),
                text_color="gray"
            ).pack(pady=50)
            return
        
        ctk.CTkLabel(
            self.resultados_frame,
            text=f"📋 EXPEDIENTES PRESTADOS ({len(resultados)})",
            font=("Segoe UI", 16, "bold"),
            text_color="#2C3E50"
        ).pack(pady=10)
        
        for r in resultados:
            nombre = f"{r[11]} {r[12]} {r[13]}".strip()
            frame = ctk.CTkFrame(self.resultados_frame, fg_color="#F8F9FA", corner_radius=8)
            frame.pack(fill="x", pady=5, padx=10)
            
            info = f"📋 #{r[10]} - {nombre}\n"
            info += f"👤 Solicitado por: {r[2]}  |  Área: {r[3]}\n"
            info += f"📅 Préstamo: {r[5]}  |  Límite: {r[6]}"
            
            ctk.CTkLabel(
                frame,
                text=info,
                font=("Segoe UI", 12),
                text_color="#333",
                justify="left",
                anchor="w"
            ).pack(pady=5, padx=10)
    
    def _reporte_disponibles(self):
        """Reporte de expedientes disponibles"""
        resultados = Expediente.obtener_todos("Disponible")
        
        if not resultados:
            ctk.CTkLabel(
                self.resultados_frame,
                text="❌ No hay expedientes disponibles.",
                font=("Segoe UI", 14),
                text_color="gray"
            ).pack(pady=50)
            return
        
        ctk.CTkLabel(
            self.resultados_frame,
            text=f"📋 EXPEDIENTES DISPONIBLES ({len(resultados)})",
            font=("Segoe UI", 16, "bold"),
            text_color="#27AE60"
        ).pack(pady=10)
        
        for r in resultados:
            nombre = f"{r[2]} {r[3]} {r[4]}".strip()
            info = f"📋 #{r[1]} - {nombre}  |  Colonia: {r[7] if r[7] else '—'}"
            ctk.CTkLabel(
                self.resultados_frame,
                text=info,
                font=("Segoe UI", 12),
                text_color="#333",
                anchor="w"
            ).pack(pady=3, padx=10)
    
    def _reporte_vencidos(self):
        """Reporte de expedientes vencidos"""
        vencidos = Prestamo.obtener_vencidos()
        
        if not vencidos:
            ctk.CTkLabel(
                self.resultados_frame,
                text="✅ No hay expedientes vencidos.",
                font=("Segoe UI", 14),
                text_color="green"
            ).pack(pady=50)
            return
        
        ctk.CTkLabel(
            self.resultados_frame,
            text=f"⏰ EXPEDIENTES VENCIDOS ({len(vencidos)})",
            font=("Segoe UI", 16, "bold"),
            text_color="#E74C3C"
        ).pack(pady=10)
        
        for r in vencidos:
            nombre = f"{r[11]} {r[12]} {r[13]}".strip()
            frame = ctk.CTkFrame(self.resultados_frame, fg_color="#FFF3CD", corner_radius=8)
            frame.pack(fill="x", pady=5, padx=10)
            
            info = f"📋 #{r[10]} - {nombre}\n"
            info += f"👤 Solicitado por: {r[2]}  |  Área: {r[3]}\n"
            info += f"⚠️ Vencido desde: {r[6]}"
            
            ctk.CTkLabel(
                frame,
                text=info,
                font=("Segoe UI", 12),
                text_color="#856404",
                justify="left",
                anchor="w"
            ).pack(pady=5, padx=10)
    
    def _reporte_prestamos_mes(self):
        """Reporte de préstamos por mes"""
        desde = self.desde.get().strip()
        hasta = self.hasta.get().strip()
        
        resultados = Prestamo.contar_por_mes(desde, hasta)
        
        if not resultados:
            ctk.CTkLabel(
                self.resultados_frame,
                text="❌ No hay préstamos en el período seleccionado.",
                font=("Segoe UI", 14),
                text_color="gray"
            ).pack(pady=50)
            return
        
        ctk.CTkLabel(
            self.resultados_frame,
            text=f"📊 PRÉSTAMOS POR MES ({desde} - {hasta})",
            font=("Segoe UI", 16, "bold"),
            text_color="#2C3E50"
        ).pack(pady=10)
        
        for r in resultados:
            info = f"📅 {r[0]}: {r[1]} préstamos"
            ctk.CTkLabel(
                self.resultados_frame,
                text=info,
                font=("Segoe UI", 13),
                text_color="#333"
            ).pack(pady=5, padx=10, anchor="w")
    
    def _reporte_areas(self):
        """Reporte de áreas más solicitantes"""
        resultados = Prestamo.contar_por_area()
        
        if not resultados:
            ctk.CTkLabel(
                self.resultados_frame,
                text="❌ No hay datos de áreas.",
                font=("Segoe UI", 14),
                text_color="gray"
            ).pack(pady=50)
            return
        
        ctk.CTkLabel(
            self.resultados_frame,
            text="🏥 ÁREAS QUE SOLICITAN MÁS EXPEDIENTES",
            font=("Segoe UI", 16, "bold"),
            text_color="#2C3E50"
        ).pack(pady=10)
        
        for r in resultados:
            info = f"🏛️ {r[0]}: {r[1]} solicitudes"
            ctk.CTkLabel(
                self.resultados_frame,
                text=info,
                font=("Segoe UI", 13),
                text_color="#333"
            ).pack(pady=5, padx=10, anchor="w")
    
    # ==========================================
    # EXPORTACIÓN SIN PANDAS
    # ==========================================
    
    def exportar_excel_sin_pandas(self):
        """Exporta a Excel usando openpyxl (sin pandas)"""
        try:
            from core.config import EXCEL_DIR
            
            resultados = Expediente.obtener_todos("Todos")
            if not resultados:
                messagebox.showwarning("Vacío", "No hay datos para exportar.")
                return
            
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Expedientes"
            
            # Encabezados
            headers = ['N° Expediente', 'Nombre', 'Apellido Paterno', 'Apellido Materno', 
                      'Fecha Nacimiento', 'Sexo', 'Colonia', 'Estado', 'Fecha Registro']
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Datos
            for row_idx, fila in enumerate(resultados, 2):
                ws.cell(row=row_idx, column=1, value=fila[1])
                ws.cell(row=row_idx, column=2, value=fila[2])
                ws.cell(row=row_idx, column=3, value=fila[3])
                ws.cell(row=row_idx, column=4, value=fila[4] if fila[4] else "")
                ws.cell(row=row_idx, column=5, value=fila[5])
                ws.cell(row=row_idx, column=6, value=fila[6])
                ws.cell(row=row_idx, column=7, value=fila[7] if fila[7] else "")
                ws.cell(row=row_idx, column=8, value=fila[8])
                ws.cell(row=row_idx, column=9, value=fila[9])
            
            # Guardar
            nombre = os.path.join(EXCEL_DIR, f"Expedientes_{datetime.now().strftime('%d%m%Y_%H%M')}.xlsx")
            wb.save(nombre)
            
            if os.name == 'nt':
                os.startfile(nombre)
            
            messagebox.showinfo("Éxito", "Excel generado correctamente.")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar: {str(e)}")
    
    def exportar_pdf_sin_pandas(self):
        """Exporta a PDF (sin pandas)"""
        try:
            from core.config import PDF_DIR
            
            resultados = Expediente.obtener_todos("Todos")
            if not resultados:
                messagebox.showwarning("Vacío", "No hay datos para exportar.")
                return
            
            pdf = ReportePDF()
            pdf.add_page()
            
            # Encabezados
            pdf.set_font("Arial", 'B', 8)
            pdf.set_fill_color(52, 73, 94)
            pdf.set_text_color(255)
            pdf.cell(25, 8, "N°", 1, 0, 'C', True)
            pdf.cell(55, 8, "NOMBRE", 1, 0, 'C', True)
            pdf.cell(30, 8, "SEXO", 1, 0, 'C', True)
            pdf.cell(35, 8, "ESTADO", 1, 0, 'C', True)
            pdf.cell(35, 8, "FECHA REG", 1, 1, 'C', True)
            
            pdf.set_font("Arial", '', 7)
            pdf.set_text_color(0)
            
            for fila in resultados:
                nombre = f"{fila[2]} {fila[3]}"[:30]
                pdf.cell(25, 6, str(fila[1]), 1)
                pdf.cell(55, 6, nombre, 1)
                pdf.cell(30, 6, fila[6], 1, 0, 'C')
                pdf.cell(35, 6, fila[8], 1, 0, 'C')
                pdf.cell(35, 6, fila[9], 1, 1, 'C')
            
            nombre = os.path.join(PDF_DIR, f"Expedientes_{datetime.now().strftime('%d%m%Y_%H%M')}.pdf")
            pdf.output(nombre)
            
            if os.name == 'nt':
                os.startfile(nombre)
            
            messagebox.showinfo("Éxito", "PDF generado correctamente.")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar: {str(e)}")